import aiosqlite
import os
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from database import DB_NAME

MONTHS_RU = [
    "", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
]


def _styles():
    return {
        "header": Font(bold=True, size=14),
        "th_font": Font(bold=True, size=10, color="FFFFFF"),
        "th_fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "total_font": Font(bold=True, size=11),
        "total_fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "worker_fill": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
        "day_fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "border": Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        ),
    }


def _cell(ws, row, col, value, s, font=None, fill=None, fmt=None, center=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = s["border"]
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if center:
        cell.alignment = Alignment(horizontal='center')
    return cell


async def generate_monthly_report(year=None, month=None):
    if year is None:
        year = date.today().year
    if month is None:
        month = date.today().month

    s = _styles()
    wb = Workbook()

    async with aiosqlite.connect(DB_NAME) as db:
        # ===== ЛИСТ 1: СВОДКА =====
        ws = wb.active
        ws.title = "Сводка"

        ws.merge_cells('A1:F1')
        ws['A1'] = f"Отчёт за {MONTHS_RU[month]} {year}"
        ws['A1'].font = s["header"]
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:F2')
        ws['A2'] = f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal='center')

        row = 4
        for col, h in enumerate(["№", "Работник", "Категории", "Записей", "Дней", "Итого (₽)"], 1):
            _cell(ws, row, col, h, s, font=s["th_font"], fill=s["th_fill"], center=True)

        cursor = await db.execute("SELECT telegram_id, name FROM workers ORDER BY name")
        workers = await cursor.fetchall()
        row = 5
        grand = 0

        for idx, (tid, name) in enumerate(workers, 1):
            cat_cursor = await db.execute("""
                SELECT c.emoji, c.name FROM worker_categories wc
                JOIN categories c ON wc.category_code = c.code WHERE wc.worker_id = ?
            """, (tid,))
            cats = await cat_cursor.fetchall()
            cats_str = ", ".join([f"{e}{n}" for e, n in cats]) if cats else "—"

            stat_cursor = await db.execute("""
                SELECT COUNT(*), COUNT(DISTINCT work_date), COALESCE(SUM(total), 0)
                FROM work_log WHERE worker_id = ?
                AND strftime('%Y', work_date) = ? AND strftime('%m', work_date) = ?
            """, (tid, str(year), f"{month:02d}"))
            cnt, days, total = await stat_cursor.fetchone()

            _cell(ws, row, 1, idx, s, center=True)
            _cell(ws, row, 2, name, s)
            _cell(ws, row, 3, cats_str, s)
            _cell(ws, row, 4, cnt, s, center=True)
            _cell(ws, row, 5, days, s, center=True)
            _cell(ws, row, 6, round(total, 2), s, fmt='#,##0.00 ₽')
            grand += total
            row += 1

        _cell(ws, row, 1, "", s, fill=s["total_fill"])
        _cell(ws, row, 2, "ИТОГО", s, font=s["total_font"], fill=s["total_fill"])
        for col in range(3, 6):
            _cell(ws, row, col, "", s, fill=s["total_fill"])
        _cell(ws, row, 6, round(grand, 2), s, font=s["total_font"],
              fill=s["total_fill"], fmt='#,##0.00 ₽')

        for col, w in zip('ABCDEF', [5, 25, 30, 12, 12, 18]):
            ws.column_dimensions[col].width = w

        # ===== ЛИСТ 2: ДЕТАЛИЗАЦИЯ =====
        ws2 = wb.create_sheet("Детализация")
        ws2.merge_cells('A1:G1')
        ws2['A1'] = f"Детализация за {MONTHS_RU[month]} {year}"
        ws2['A1'].font = s["header"]
        ws2['A1'].alignment = Alignment(horizontal='center')

        row = 3
        for tid, name in workers:
            ws2.merge_cells(f'A{row}:G{row}')
            cell = ws2.cell(row=row, column=1, value=f"👤 {name}")
            cell.font = Font(bold=True, size=12)
            cell.fill = s["worker_fill"]
            row += 1

            for col, h in enumerate(["Дата", "Работа", "Категория", "Кол-во",
                                       "Расценка", "Сумма", "Время"], 1):
                _cell(ws2, row, col, h, s, font=s["th_font"], fill=s["th_fill"], center=True)
            row += 1

            rec_cursor = await db.execute("""
                SELECT wl.work_date, pl.name, c.name, wl.quantity,
                       wl.price_per_unit, wl.total, wl.created_at
                FROM work_log wl
                JOIN price_list pl ON wl.work_code = pl.code
                JOIN categories c ON pl.category_code = c.code
                WHERE wl.worker_id = ?
                  AND strftime('%Y', wl.work_date) = ?
                  AND strftime('%m', wl.work_date) = ?
                ORDER BY wl.work_date, wl.created_at
            """, (tid, str(year), f"{month:02d}"))
            records = await rec_cursor.fetchall()

            wtotal = 0
            cur_date = ""
            day_total = 0

            for rec in records:
                if rec[0] != cur_date and cur_date != "":
                    for c2 in range(1, 6):
                        _cell(ws2, row, c2, "", s, fill=s["day_fill"])
                    _cell(ws2, row, 5, f"День {cur_date}:", s,
                          font=Font(bold=True, italic=True, size=9), fill=s["day_fill"])
                    _cell(ws2, row, 6, round(day_total, 2), s,
                          font=Font(bold=True, italic=True, size=9),
                          fill=s["day_fill"], fmt='#,##0.00')
                    _cell(ws2, row, 7, "", s, fill=s["day_fill"])
                    row += 1
                    day_total = 0
                cur_date = rec[0]

                _cell(ws2, row, 1, rec[0], s)
                _cell(ws2, row, 2, rec[1], s)
                _cell(ws2, row, 3, rec[2], s)
                _cell(ws2, row, 4, rec[3], s, center=True)
                _cell(ws2, row, 5, round(rec[4], 2), s, fmt='#,##0.00')
                _cell(ws2, row, 6, round(rec[5], 2), s, fmt='#,##0.00')
                _cell(ws2, row, 7, rec[6], s)
                wtotal += rec[5]
                day_total += rec[5]
                row += 1

            if cur_date != "":
                for c2 in range(1, 6):
                    _cell(ws2, row, c2, "", s, fill=s["day_fill"])
                _cell(ws2, row, 5, f"День {cur_date}:", s,
                      font=Font(bold=True, italic=True, size=9), fill=s["day_fill"])
                _cell(ws2, row, 6, round(day_total, 2), s,
                      font=Font(bold=True, italic=True, size=9),
                      fill=s["day_fill"], fmt='#,##0.00')
                _cell(ws2, row, 7, "", s, fill=s["day_fill"])
                row += 1

            if records:
                for col in range(1, 6):
                    _cell(ws2, row, col, "", s, fill=s["total_fill"])
                _cell(ws2, row, 2, f"ИТОГО {name}:", s, font=s["total_font"], fill=s["total_fill"])
                _cell(ws2, row, 6, round(wtotal, 2), s, font=s["total_font"],
                      fill=s["total_fill"], fmt='#,##0.00 ₽')
                _cell(ws2, row, 7, "", s, fill=s["total_fill"])
                row += 1
            else:
                ws2.cell(row=row, column=1, value="Нет записей")
                row += 1
            row += 1

        for col, w in zip('ABCDEFG', [14, 25, 20, 10, 14, 14, 20]):
            ws2.column_dimensions[col].width = w

        # ===== ЛИСТ 3: ПО ДНЯМ =====
        ws3 = wb.create_sheet("По дням")
        ws3.merge_cells('A1:D1')
        ws3['A1'] = f"По дням за {MONTHS_RU[month]} {year}"
        ws3['A1'].font = s["header"]
        ws3['A1'].alignment = Alignment(horizontal='center')

        row = 3
        for col, h in enumerate(["Дата", "Работник", "Работа", "Сумма (₽)"], 1):
            _cell(ws3, row, col, h, s, font=s["th_font"], fill=s["th_fill"], center=True)
        row += 1

        daily_cursor = await db.execute("""
            SELECT wl.work_date, w.name, pl.name, SUM(wl.total)
            FROM work_log wl
            JOIN workers w ON wl.worker_id = w.telegram_id
            JOIN price_list pl ON wl.work_code = pl.code
            WHERE strftime('%Y', wl.work_date) = ? AND strftime('%m', wl.work_date) = ?
            GROUP BY wl.work_date, w.name, pl.name
            ORDER BY wl.work_date, w.name
        """, (str(year), f"{month:02d}"))
        daily = await daily_cursor.fetchall()

        cur_date = ""
        day_sum = 0
        for wd, wn, wname, total in daily:
            if wd != cur_date and cur_date != "":
                _cell(ws3, row, 1, "", s)
                _cell(ws3, row, 2, f"Итого за {cur_date}:", s,
                      font=Font(bold=True, italic=True, size=9))
                _cell(ws3, row, 3, "", s)
                _cell(ws3, row, 4, round(day_sum, 2), s,
                      font=Font(bold=True, size=9), fmt='#,##0.00')
                row += 1
                day_sum = 0
            cur_date = wd
            _cell(ws3, row, 1, wd, s)
            _cell(ws3, row, 2, wn, s)
            _cell(ws3, row, 3, wname, s)
            _cell(ws3, row, 4, round(total, 2), s, fmt='#,##0.00')
            day_sum += total
            row += 1

        if cur_date:
            _cell(ws3, row, 1, "", s)
            _cell(ws3, row, 2, f"Итого за {cur_date}:", s,
                  font=Font(bold=True, italic=True, size=9))
            _cell(ws3, row, 3, "", s)
            _cell(ws3, row, 4, round(day_sum, 2), s,
                  font=Font(bold=True, size=9), fmt='#,##0.00')

        for col, w in zip('ABCD', [14, 25, 25, 15]):
            ws3.column_dimensions[col].width = w

        # ===== ЛИСТ 4: ПО КАТЕГОРИЯМ =====
        ws4 = wb.create_sheet("По категориям")
        ws4.merge_cells('A1:E1')
        ws4['A1'] = f"По категориям за {MONTHS_RU[month]} {year}"
        ws4['A1'].font = s["header"]
        ws4['A1'].alignment = Alignment(horizontal='center')

        row = 3
        for col, h in enumerate(["Категория", "Работа", "Кол-во", "Расценка", "Итого (₽)"], 1):
            _cell(ws4, row, col, h, s, font=s["th_font"], fill=s["th_fill"], center=True)
        row += 1

        cat_cursor = await db.execute("""
            SELECT c.name, pl.name, SUM(wl.quantity), wl.price_per_unit, SUM(wl.total)
            FROM work_log wl
            JOIN price_list pl ON wl.work_code = pl.code
            JOIN categories c ON pl.category_code = c.code
            WHERE strftime('%Y', wl.work_date) = ? AND strftime('%m', wl.work_date) = ?
            GROUP BY c.name, pl.name, wl.price_per_unit
            ORDER BY c.name, pl.name
        """, (str(year), f"{month:02d}"))
        cat_data = await cat_cursor.fetchall()

        cat_grand = 0
        for cn, pn, qty, price, total in cat_data:
            _cell(ws4, row, 1, cn, s)
            _cell(ws4, row, 2, pn, s)
            _cell(ws4, row, 3, qty, s, center=True)
            _cell(ws4, row, 4, round(price, 2), s, fmt='#,##0.00')
            _cell(ws4, row, 5, round(total, 2), s, fmt='#,##0.00')
            cat_grand += total
            row += 1

        for col in range(1, 5):
            _cell(ws4, row, col, "", s, fill=s["total_fill"])
        _cell(ws4, row, 2, "ОБЩИЙ ИТОГО", s, font=s["total_font"], fill=s["total_fill"])
        _cell(ws4, row, 5, round(cat_grand, 2), s, font=s["total_font"],
              fill=s["total_fill"], fmt='#,##0.00 ₽')

        for col, w in zip('ABCDE', [20, 25, 12, 14, 15]):
            ws4.column_dimensions[col].width = w

        # ===== ЛИСТ 5: БАЛАНС =====
        ws5 = wb.create_sheet("Баланс")
        ws5.merge_cells('A1:G1')
        ws5['A1'] = f"Баланс за {MONTHS_RU[month]} {year}"
        ws5['A1'].font = s["header"]
        ws5['A1'].alignment = Alignment(horizontal='center')

        row = 3
        for col, h in enumerate(["№", "Работник", "Заработано", "Авансы",
                                   "Штрафы", "К выплате", "Дней"], 1):
            _cell(ws5, row, col, h, s, font=s["th_font"], fill=s["th_fill"], center=True)
        row += 1

        balance_cursor = await db.execute("""
            SELECT 
                w.telegram_id, w.name,
                COALESCE(earn.total_earned, 0) as earned,
                COALESCE(adv.total_advance, 0) as advances,
                COALESCE(pen.total_penalty, 0) as penalties,
                COALESCE(earn.work_days, 0) as work_days
            FROM workers w
            LEFT JOIN (
                SELECT worker_id, 
                       SUM(total) as total_earned,
                       COUNT(DISTINCT work_date) as work_days
                FROM work_log
                WHERE strftime('%Y', work_date) = ? AND strftime('%m', work_date) = ?
                GROUP BY worker_id
            ) earn ON w.telegram_id = earn.worker_id
            LEFT JOIN (
                SELECT worker_id, SUM(amount) as total_advance
                FROM advances
                WHERE strftime('%Y', advance_date) = ? AND strftime('%m', advance_date) = ?
                GROUP BY worker_id
            ) adv ON w.telegram_id = adv.worker_id
            LEFT JOIN (
                SELECT worker_id, SUM(amount) as total_penalty
                FROM penalties
                WHERE strftime('%Y', penalty_date) = ? AND strftime('%m', penalty_date) = ?
                GROUP BY worker_id
            ) pen ON w.telegram_id = pen.worker_id
            ORDER BY w.name
        """, (str(year), f"{month:02d}", str(year), f"{month:02d}", str(year), f"{month:02d}"))
        balance_data = await balance_cursor.fetchall()

        grand_earned = 0
        grand_advances = 0
        grand_penalties = 0
        grand_to_pay = 0
        idx = 0

        for tid, name, earned, advances, penalties, work_days in balance_data:
            if earned == 0 and advances == 0 and penalties == 0:
                continue
            idx += 1
            to_pay = earned - advances - penalties

            _cell(ws5, row, 1, idx, s, center=True)
            _cell(ws5, row, 2, name, s)
            _cell(ws5, row, 3, round(earned, 2), s, fmt='#,##0.00 ₽')
            _cell(ws5, row, 4, round(advances, 2), s, fmt='#,##0.00 ₽')
            _cell(ws5, row, 5, round(penalties, 2), s, fmt='#,##0.00 ₽')
            _cell(ws5, row, 6, round(to_pay, 2), s, fmt='#,##0.00 ₽')
            _cell(ws5, row, 7, work_days, s, center=True)

            grand_earned += earned
            grand_advances += advances
            grand_penalties += penalties
            grand_to_pay += to_pay
            row += 1

        _cell(ws5, row, 1, "", s, fill=s["total_fill"])
        _cell(ws5, row, 2, "ИТОГО", s, font=s["total_font"], fill=s["total_fill"])
        _cell(ws5, row, 3, round(grand_earned, 2), s, font=s["total_font"],
              fill=s["total_fill"], fmt='#,##0.00 ₽')
        _cell(ws5, row, 4, round(grand_advances, 2), s, font=s["total_font"],
              fill=s["total_fill"], fmt='#,##0.00 ₽')
        _cell(ws5, row, 5, round(grand_penalties, 2), s, font=s["total_font"],
              fill=s["total_fill"], fmt='#,##0.00 ₽')
        _cell(ws5, row, 6, round(grand_to_pay, 2), s, font=s["total_font"],
              fill=s["total_fill"], fmt='#,##0.00 ₽')
        _cell(ws5, row, 7, "", s, fill=s["total_fill"])

        row += 2

        # Детализация авансов
        adv_detail_cursor = await db.execute("""
            SELECT w.name, a.advance_date, a.amount, a.comment
            FROM advances a
            JOIN workers w ON a.worker_id = w.telegram_id
            WHERE strftime('%Y', a.advance_date) = ? AND strftime('%m', a.advance_date) = ?
            ORDER BY w.name, a.advance_date
        """, (str(year), f"{month:02d}"))
        adv_details = await adv_detail_cursor.fetchall()

        if adv_details:
            ws5.merge_cells(f'A{row}:G{row}')
            cell = ws5.cell(row=row, column=1, value="Детализация авансов")
            cell.font = Font(bold=True, size=12)
            cell.fill = s["worker_fill"]
            row += 1

            for col, h in enumerate(["Работник", "Дата", "Сумма", "Комментарий"], 1):
                _cell(ws5, row, col, h, s, font=s["th_font"], fill=s["th_fill"], center=True)
            row += 1

            for name, adv_date, amount, comment in adv_details:
                _cell(ws5, row, 1, name, s)
                _cell(ws5, row, 2, adv_date, s)
                _cell(ws5, row, 3, round(amount, 2), s, fmt='#,##0.00 ₽')
                _cell(ws5, row, 4, comment or "", s)
                row += 1

        row += 1

        # Детализация штрафов
        pen_detail_cursor = await db.execute("""
            SELECT w.name, p.penalty_date, p.amount, p.reason
            FROM penalties p
            JOIN workers w ON p.worker_id = w.telegram_id
            WHERE strftime('%Y', p.penalty_date) = ? AND strftime('%m', p.penalty_date) = ?
            ORDER BY w.name, p.penalty_date
        """, (str(year), f"{month:02d}"))
        pen_details = await pen_detail_cursor.fetchall()

        if pen_details:
            ws5.merge_cells(f'A{row}:G{row}')
            cell = ws5.cell(row=row, column=1, value="Детализация штрафов")
            cell.font = Font(bold=True, size=12)
            cell.fill = s["worker_fill"]
            row += 1

            for col, h in enumerate(["Работник", "Дата", "Сумма", "Причина"], 1):
                _cell(ws5, row, col, h, s, font=s["th_font"], fill=s["th_fill"], center=True)
            row += 1

            for name, pen_date, amount, reason in pen_details:
                _cell(ws5, row, 1, name, s)
                _cell(ws5, row, 2, pen_date, s)
                _cell(ws5, row, 3, round(amount, 2), s, fmt='#,##0.00 ₽')
                _cell(ws5, row, 4, reason or "", s)
                row += 1

        for col, w in zip('ABCDEFG', [5, 25, 18, 18, 18, 18, 10]):
            ws5.column_dimensions[col].width = w

    filename = f"report_{year}_{month:02d}.xlsx"
    wb.save(filename)
    return filename


async def generate_worker_report(worker_id, worker_name, year=None, month=None):
    if year is None:
        year = date.today().year
    if month is None:
        month = date.today().month

    s = _styles()
    wb = Workbook()
    ws = wb.active
    ws.title = f"Отчёт {worker_name}"

    ws.merge_cells('A1:F1')
    ws['A1'] = f"Отчёт: {worker_name}"
    ws['A1'].font = s["header"]
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = f"Период: {MONTHS_RU[month]} {year}"
    ws['A2'].alignment = Alignment(horizontal='center')

    row = 4
    for col, h in enumerate(["Дата", "Категория", "Работа", "Кол-во", "Расценка", "Сумма"], 1):
        _cell(ws, row, col, h, s, font=s["th_font"], fill=s["th_fill"], center=True)
    row += 1

    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("""
            SELECT wl.work_date, c.name, pl.name, wl.quantity, wl.price_per_unit, wl.total
            FROM work_log wl
            JOIN price_list pl ON wl.work_code = pl.code
            JOIN categories c ON pl.category_code = c.code
            WHERE wl.worker_id = ?
              AND strftime('%Y', wl.work_date) = ?
              AND strftime('%m', wl.work_date) = ?
            ORDER BY wl.work_date, wl.created_at
        """, (worker_id, str(year), f"{month:02d}"))
        records = await cursor.fetchall()

    grand = 0
    cur_date = ""
    day_total = 0

    for wd, cn, pn, qty, price, total in records:
        if wd != cur_date and cur_date != "":
            for c2 in range(1, 5):
                _cell(ws, row, c2, "", s, fill=s["day_fill"])
            _cell(ws, row, 4, f"День {cur_date}:", s,
                  font=Font(bold=True, italic=True, size=9), fill=s["day_fill"])
            _cell(ws, row, 5, "", s, fill=s["day_fill"])
            _cell(ws, row, 6, round(day_total, 2), s,
                  font=Font(bold=True, italic=True, size=9),
                  fill=s["day_fill"], fmt='#,##0.00')
            row += 1
            day_total = 0
        cur_date = wd

        _cell(ws, row, 1, wd, s)
        _cell(ws, row, 2, cn, s)
        _cell(ws, row, 3, pn, s)
        _cell(ws, row, 4, qty, s, center=True)
        _cell(ws, row, 5, round(price, 2), s, fmt='#,##0.00')
        _cell(ws, row, 6, round(total, 2), s, fmt='#,##0.00')
        grand += total
        day_total += total
        row += 1

    if cur_date != "":
        for c2 in range(1, 5):
            _cell(ws, row, c2, "", s, fill=s["day_fill"])
        _cell(ws, row, 4, f"День {cur_date}:", s,
              font=Font(bold=True, italic=True, size=9), fill=s["day_fill"])
        _cell(ws, row, 5, "", s, fill=s["day_fill"])
        _cell(ws, row, 6, round(day_total, 2), s,
              font=Font(bold=True, italic=True, size=9),
              fill=s["day_fill"], fmt='#,##0.00')
        row += 1

    for col in range(1, 6):
        _cell(ws, row, col, "", s, fill=s["total_fill"])
    _cell(ws, row, 3, "ИТОГО:", s, font=s["total_font"], fill=s["total_fill"])
    _cell(ws, row, 6, round(grand, 2), s, font=s["total_font"],
          fill=s["total_fill"], fmt='#,##0.00 ₽')
    row += 1

    # ===== БАЛАНС =====
    row += 1

    async with aiosqlite.connect(DB_NAME) as db:
        adv_cursor = await db.execute("""
            SELECT advance_date, amount, comment FROM advances
            WHERE worker_id = ? AND strftime('%Y', advance_date) = ? AND strftime('%m', advance_date) = ?
            ORDER BY advance_date
        """, (worker_id, str(year), f"{month:02d}"))
        advances = await adv_cursor.fetchall()

        pen_cursor = await db.execute("""
            SELECT penalty_date, amount, reason FROM penalties
            WHERE worker_id = ? AND strftime('%Y', penalty_date) = ? AND strftime('%m', penalty_date) = ?
            ORDER BY penalty_date
        """, (worker_id, str(year), f"{month:02d}"))
        penalties = await pen_cursor.fetchall()

    total_advances = sum(a[1] for a in advances)
    total_penalties = sum(p[1] for p in penalties)
    to_pay = grand - total_advances - total_penalties

    ws.merge_cells(f'A{row}:F{row}')
    cell = ws.cell(row=row, column=1, value="БАЛАНС")
    cell.font = Font(bold=True, size=12)
    cell.fill = s["worker_fill"]
    row += 1

    balance_items = [
        ("Заработано:", grand),
        ("Авансы:", total_advances),
        ("Штрафы:", total_penalties),
        ("К выплате:", to_pay),
    ]

    for label, value in balance_items:
        _cell(ws, row, 3, label, s, font=Font(bold=True))
        _cell(ws, row, 6, round(value, 2), s, font=Font(bold=True), fmt='#,##0.00 ₽')
        row += 1

    if advances:
        row += 1
        ws.cell(row=row, column=1, value="Авансы:").font = Font(bold=True, size=11)
        row += 1
        for adv_date, amount, comment in advances:
            _cell(ws, row, 1, adv_date, s)
            _cell(ws, row, 3, comment or "", s)
            _cell(ws, row, 6, round(amount, 2), s, fmt='#,##0.00 ₽')
            row += 1

    if penalties:
        row += 1
        ws.cell(row=row, column=1, value="Штрафы:").font = Font(bold=True, size=11)
        row += 1
        for pen_date, amount, reason in penalties:
            _cell(ws, row, 1, pen_date, s)
            _cell(ws, row, 3, reason or "", s)
            _cell(ws, row, 6, round(amount, 2), s, fmt='#,##0.00 ₽')
            row += 1

    for col, w in zip('ABCDEF', [14, 20, 25, 10, 14, 14]):
        ws.column_dimensions[col].width = w

    filename = f"report_{worker_name}_{year}_{month:02d}.xlsx"
    wb.save(filename)
    return filename