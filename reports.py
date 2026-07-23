import asyncpg
import os
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

DATABASE_URL = os.getenv("DATABASE_URL", "")

MONTHS_RU = [
    "", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
]


def _styles():
    return {
        "header": Font(bold=True, size=14),
        "th_font": Font(bold=True, size=10, color="FFFFFF"),
        "th_fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "total_font": Font(bold=True, size=11),
        "total_fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "worker_fill": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
        "day_fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "border": Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        ),
    }


def _cell(ws, row, col, value, s, font=None, fill=None, fmt=None, center=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = s["border"]
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if center:
        cell.alignment = Alignment(horizontal='center')
    return cell


async def generate_monthly_report(year=None, month=None):
    if year is None:
        year = date.today().year
    if month is None:
        month = date.today().month

    s = _styles()
    wb = Workbook()

    conn = await asyncpg.connect(DATABASE_URL)
    try:
        # ===== ЛИСТ 1: СВОДКА =====
        ws = wb.active
        ws.title = "Сводка"

        ws.merge_cells('A1:F1')
        ws['A1'] = f"Отчёт за {MONTHS_RU[month]} {year}"
        ws['A1'].font = s["header"]
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:F2')
        ws['A2'] = f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal='center')

        row = 4
        for col, h in enumerate(["№", "Работник", "Категории", "Записей", "Дней", "Итого (₽)"], 1):
            _cell(ws, row, col, h, s, font=s["th_font"], fill=s["th_fill"], center=True)

        workers = await conn.fetch("SELECT telegram_id, name FROM workers ORDER BY name")
        row = 5
        grand = 0

        for idx, worker in enumerate(workers, 1):
            tid, name = worker['telegram_id'], worker['name']
            
            cats = await conn.fetch("""
                SELECT c.emoji, c.name FROM worker_categories wc
                JOIN categories c ON wc.category_code = c.code WHERE wc.worker_id = $1
            """, tid)
            cats_str = ", ".join([f"{c['emoji']}{c['name']}" for c in cats]) if cats else "—"

            stat = await conn.fetchrow("""
                SELECT COUNT(*), COUNT(DISTINCT work_date), COALESCE(SUM(total), 0)
                FROM work_log WHERE worker_id = $1
                AND EXTRACT(YEAR FROM work_date) = $2 AND EXTRACT(MONTH FROM work_date) = $3
            """, tid, year, month)
            cnt, days, total = stat[0], stat[1], stat[2]

            _cell(ws, row, 1, idx, s, center=True)
            _cell(ws, row, 2, name, s)
            _cell(ws, row, 3, cats_str, s)
            _cell(ws, row, 4, cnt, s, center=True)
            _cell(ws, row, 5, days, s, center=True)
            _cell(ws, row, 6, round(total, 2), s, fmt='#,##0.00 ₽')
            grand += total
            row += 1

        _cell(ws, row, 1, "", s, fill=s["total_fill"])
        _cell(ws, row, 2, "ИТОГО", s, font=s["total_font"], fill=s["total_fill"])
        for col in range(3, 6):
            _cell(ws, row, col, "", s, fill=s["total_fill"])
        _cell(ws, row, 6, round(grand, 2), s, font=s["total_font"],
              fill=s["total_fill"], fmt='#,##0.00 ₽')

        for col, w in zip('ABCDEF', [5, 25, 30, 12, 12, 18]):
            ws.column_dimensions[col].width = w

        # ===== ЛИСТ 2: ДЕТАЛИЗАЦИЯ =====
        ws2 = wb.create_sheet("Детализация")
        ws2.merge_cells('A1:G1')
        ws2['A1'] = f"Детализация за {MONTHS_RU[month]} {year}"
        ws2['A1'].font = s["header"]
        ws2['A1'].alignment = Alignment(horizontal='center')

        row = 3
        for worker in workers:
            tid, name = worker['telegram_id'], worker['name']
            
            ws2.merge_cells(f'A{row}:G{row}')
            cell = ws2.cell(row=row, column=1, value=f"👤 {name}")
            cell.font = Font(bold=True, size=12)
            cell.fill = s["worker_fill"]
            row += 1

            for col, h in enumerate(["Дата", "Работа", "Категория", "Кол-во",
                                       "Расценка", "Сумма", "Время"], 1):
                _cell(ws2, row, col, h, s, font=s["th_font"], fill=s["th_fill"], center=True)
            row += 1

            records = await conn.fetch("""
                SELECT wl.work_date::TEXT, pl.name, c.name, wl.quantity,
                       wl.price_per_unit, wl.total, wl.created_at::TEXT
                FROM work_log wl
                JOIN price_list pl ON wl.work_code = pl.code
                JOIN categories c ON pl.category_code = c.code
                WHERE wl.worker_id = $1
                  AND EXTRACT(YEAR FROM wl.work_date) = $2
                  AND EXTRACT(MONTH FROM wl.work_date) = $3
                ORDER BY wl.work_date, wl.created_at
            """, tid, year, month)

            wtotal = 0
            cur_date = ""
            day_total = 0

            for rec in records:
                if rec[0] != cur_date and cur_date != "":
                    for c2 in range(1, 6):
                        _cell(ws2, row, c2, "", s, fill=s["day_fill"])
                    _cell(ws2, row, 5, f"День {cur_date}:", s,
                          font=Font(bold=True, italic=True, size=9), fill=s["day_fill"])
                    _cell(ws2, row, 6, round(day_total, 2), s,
                          font=Font(bold=True, italic=True, size=9),
                          fill=s["day_fill"], fmt='#,##0.00')
                    _cell(ws2, row, 7, "", s, fill=s["day_fill"])
                    row += 1
                    day_total = 0
                cur_date = rec[0]

                _cell(ws2, row, 1, rec[0], s)
                _cell(ws2, row, 2, rec[1], s)
                _cell(ws2, row, 3, rec[2], s)
                _cell(ws2, row, 4, rec[3], s, center=True)
                _cell(ws2, row, 5, round(rec[4], 2), s, fmt='#,##0.00')
                _cell(ws2, row, 6, round(rec[5], 2), s, fmt='#,##0.00')
                _cell(ws2, row, 7, rec[6], s)
                wtotal += rec[5]
                day_total += rec[5]
                row += 1

            if cur_date != "":
                for c2 in range(1, 6):
                    _cell(ws2, row, c2, "", s, fill=s["day_fill"])
                _cell(ws2, row, 5, f"День {cur_date}:", s,
                      font=Font(bold=True, italic=True, size=9), fill=s["day_fill"])
                _cell(ws2, row, 6, round(day_total, 2), s,
                      font=Font(bold=True, italic=True, size=9),
                      fill=s["day_fill"], fmt='#,##0.00')
                _cell(ws2, row, 7, "", s, fill=s["day_fill"])
                row += 1

            if records:
                for col in range(1, 6):
                    _cell(ws2, row, col, "", s, fill=s["total_fill"])
                _cell(ws2, row, 2, f"ИТОГО {name}:", s, font=s["total_font"], fill=s["total_fill"])
                _cell(ws2, row, 6, round(wtotal, 2), s, font=s["total_font"],
                      fill=s["total_fill"], fmt='#,##0.00 ₽')
                _cell(ws2, row, 7, "", s, fill=s["total_fill"])
                row += 1
            else:
                ws2.cell(row=row, column=1, value="Нет записей")
                row += 1
            row += 1

        for col, w in zip('ABCDEFG', [14, 25, 20, 10, 14, 14, 20]):
            ws2.column_dimensions[col].width = w

        # ===== ЛИСТ 3: ПО ДНЯМ =====
        ws3 = wb.create_sheet("По дням")
        ws3.merge_cells('A1:D1')
        ws3['A1'] = f"По дням за {MONTHS_RU[month]} {year}"
        ws3['A1'].font = s["header"]
        ws3['A1'].alignment = Alignment(horizontal='center')

        row = 3
        for col, h in enumerate(["Дата", "Работник", "Работа", "Сумма (₽)"], 1):
            _cell(ws3, row, col, h, s, font=s["th_font"], fill=s["th_fill"], center=True)
        row += 1

        daily = await conn.fetch("""
            SELECT wl.work_date::TEXT, w.name, pl.name, SUM(wl.total)
            FROM work_log wl
            JOIN workers w ON wl.worker_id = w.telegram_id
            JOIN price_list pl ON wl.work_code = pl.code
            WHERE EXTRACT(YEAR FROM wl.work_date) = $1 AND EXTRACT(MONTH FROM wl.work_date) = $2
            GROUP BY wl.work_date, w.name, pl.name
            ORDER BY wl.work_date, w.name
        """, year, month)

        cur_date = ""
        day_sum = 0
        for rec in daily:
            wd, wn, wname, total = rec[0], rec[1], rec[2], rec[3]
            if wd != cur_date and cur_date != "":
                _cell(ws3, row, 1, "", s)
                _cell(ws3, row, 2, f"Итого за {cur_date}:", s,
                      font=Font(bold=True, italic=True, size=9))
                _cell(ws3, row, 3, "", s)
                _cell(ws3, row, 4, round(day_sum, 2), s,
                      font=Font(bold=True, size=9), fmt='#,##0.00')
                row += 1
                day_sum = 0
            cur_date = wd
            _cell(ws3, row, 1, wd, s)
            _cell(ws3, row, 2, wn, s)
            _cell(ws3, row, 3, wname, s)
            _cell(ws3, row, 4, round(total, 2), s, fmt='#,##0.00')
            day_sum += total
            row += 1

        if cur_date:
            _cell(ws3, row, 1, "", s)
            _cell(ws3, row, 2, f"Итого за {cur_date}:", s,
                  font=Font(bold=True, italic=True, size=9))
            _cell(ws3, row, 3, "", s)
            _cell(ws3, row, 4, round(day_sum, 2), s,
                  font=Font(bold=True, size=9), fmt='#,##0.00')

        for col, w in zip('ABCD', [14, 25, 25, 15]):
            ws3.column_dimensions[col].width = w

        # ===== ЛИСТ 4: ПО КАТЕГОРИЯМ =====
        ws4 = wb.create_sheet("По категориям")
        ws4.merge_cells('A1:E1')
        ws4['A1'] = f"По категориям за {MONTHS_RU[month]} {year}"
        ws4['A1'].font = s["header"]
        ws4['A1'].alignment = Alignment(horizontal='center')

        row = 3
        for col, h in enumerate(["Категория", "Работа", "Кол-во", "Расценка", "Итого (₽)"], 1):
            _cell(ws4, row, col, h, s, font=s["th_font"], fill=s["th_fill"], center=True)
        row += 1

        cat_data = await conn.fetch("""
            SELECT c.name, pl.name, SUM(wl.quantity), wl.price_per_unit, SUM(wl.total)
            FROM work_log wl
            JOIN price_list pl ON wl.work_code = pl.code
            JOIN categories c ON pl.category_code = c.code
            WHERE EXTRACT(YEAR FROM wl.work_date) = $1 AND EXTRACT(MONTH FROM wl.work_date) = $2
            GROUP BY c.name, pl.name, wl.price_per_unit
            ORDER BY c.name, pl.name
        """, year, month)

        cat_grand = 0
        for rec in cat_data:
            cn, pn, qty, price, total = rec[0], rec[1], rec[2], rec[3], rec[4]
            _cell(ws4, row, 1, cn, s)
            _cell(ws4, row, 2, pn, s)
            _cell(ws4, row, 3, qty, s, center=True)
            _cell(ws4, row, 4, round(price, 2), s, fmt='#,##0.00')
            _cell(ws4, row, 5, round(total, 2), s, fmt='#,##0.00')
            cat_grand += total
            row += 1

        for col in range(1, 5):
            _cell(ws4, row, col, "", s, fill=s["total_fill"])
        _cell(ws4, row, 2, "ОБЩИЙ ИТОГО", s, font=s["total_font"], fill=s["total_fill"])
        _cell(ws4, row, 5, round(cat_grand, 2), s, font=s["total_font"],
              fill=s["total_fill"], fmt='#,##0.00 ₽')

        for col, w in zip('ABCDE', [20, 25, 12, 14, 15]):
            ws4.column_dimensions[col].width = w

    finally:
        await conn.close()

    filename = f"report_{year}_{month:02d}.xlsx"
    wb.save(filename)
    return filename


async def generate_worker_report(worker_id, worker_name, year=None, month=None):
    if year is None:
        year = date.today().year
    if month is None:
        month = date.today().month

    s = _styles()
    wb = Workbook()
    ws = wb.active
    ws.title = f"Отчёт {worker_name}"

    ws.merge_cells('A1:F1')
    ws['A1'] = f"Отчёт: {worker_name}"
    ws['A1'].font = s["header"]
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = f"Период: {MONTHS_RU[month]} {year}"
    ws['A2'].alignment = Alignment(horizontal='center')

    row = 4
    for col, h in enumerate(["Дата", "Категория", "Работа", "Кол-во", "Расценка", "Сумма"], 1):
        _cell(ws, row, col, h, s, font=s["th_font"], fill=s["th_fill"], center=True)
    row += 1

    conn = await asyncpg.connect(DATABASE_URL)
    try:
        records = await conn.fetch("""
            SELECT wl.work_date::TEXT, c.name, pl.name, wl.quantity, wl.price_per_unit, wl.total
            FROM work_log wl
            JOIN price_list pl ON wl.work_code = pl.code
            JOIN categories c ON pl.category_code = c.code
            WHERE wl.worker_id = $1
              AND EXTRACT(YEAR FROM wl.work_date) = $2
              AND EXTRACT(MONTH FROM wl.work_date) = $3
            ORDER BY wl.work_date, wl.created_at
        """, worker_id, year, month)
    finally:
        await conn.close()

    grand = 0
    cur_date = ""
    day_total = 0

    for rec in records:
        wd, cn, pn, qty, price, total = rec[0], rec[1], rec[2], rec[3], rec[4], rec[5]
        if wd != cur_date and cur_date != "":
            for c2 in range(1, 5):
                _cell(ws, row, c2, "", s, fill=s["day_fill"])
            _cell(ws, row, 4, f"День {cur_date}:", s,
                  font=Font(bold=True, italic=True, size=9), fill=s["day_fill"])
            _cell(ws, row, 5, "", s, fill=s["day_fill"])
            _cell(ws, row, 6, round(day_total, 2), s,
                  font=Font(bold=True, italic=True, size=9),
                  fill=s["day_fill"], fmt='#,##0.00')
            row += 1
            day_total = 0
        cur_date = wd

        _cell(ws, row, 1, wd, s)
        _cell(ws, row, 2, cn, s)
        _cell(ws, row, 3, pn, s)
        _cell(ws, row, 4, qty, s, center=True)
        _cell(ws, row, 5, round(price, 2), s, fmt='#,##0.00')
        _cell(ws, row, 6, round(total, 2), s, fmt='#,##0.00')
        grand += total
        day_total += total
        row += 1

    if cur_date != "":
        for c2 in range(1, 5):
            _cell(ws, row, c2, "", s, fill=s["day_fill"])
        _cell(ws, row, 4, f"День {cur_date}:", s,
              font=Font(bold=True, italic=True, size=9), fill=s["day_fill"])
        _cell(ws, row, 5, "", s, fill=s["day_fill"])
        _cell(ws, row, 6, round(day_total, 2), s,
              font=Font(bold=True, italic=True, size=9),
              fill=s["day_fill"], fmt='#,##0.00')
        row += 1

    for col in range(1, 6):
        _cell(ws, row, col, "", s, fill=s["total_fill"])
    _cell(ws, row, 3, "ИТОГО:", s, font=s["total_font"], fill=s["total_fill"])
    _cell(ws, row, 6, round(grand, 2), s, font=s["total_font"],
          fill=s["total_fill"], fmt='#,##0.00 ₽')

    for col, w in zip('ABCDEF', [14, 20, 25, 10, 14, 14]):
        ws.column_dimensions[col].width = w

    filename = f"report_{worker_name}_{year}_{month:02d}.xlsx"
    wb.save(filename)
    return filename


async def generate_salary_report(year: int, month: int, workers_data: list) -> str:
    """
    Ведомость выплаты зарплаты + сверка кассы.
    workers_data: список dict с ключами name, work_summary, fixed_salary,
                  bonus, penalties, advances, balance

    Колонки:
      A Работник | B Что сделал | C Заработано | D Фикс.ставка | E Премия |
      F Штрафы | G Авансы | H Надо выплатить (баланс) | I Наличными |
      J Перевод | K Выплачено (=I+J) | L Остаток (=H-K)

    Графы «Наличными» и «Перевод» заполняются вручную — «Выплачено», «Остаток»,
    итоги и сверка кассы пересчитываются формулами Excel автоматически.
    """
    s = _styles()
    money_fmt = '#,##0.00 ₽'
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # жёлтый — «заполнить вручную»
    warn_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")   # для недостачи
    tf = s["total_fill"]
    tf_font = s["total_font"]

    wb = Workbook()
    ws = wb.active
    ws.title = f"Ведомость {MONTHS_RU[month]} {year}"

    # ---- Заголовок ----
    ws.merge_cells('A1:L1')
    ws['A1'] = f"Ведомость зарплаты — {MONTHS_RU[month]} {year}"
    ws['A1'].font = s["header"]
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:L2')
    ws['A2'] = f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}  •  жёлтые ячейки — заполнить вручную"
    ws['A2'].alignment = Alignment(horizontal='center')

    # ---- Шапка таблицы ----
    headers = ["Работник", "Что сделал", "Заработано", "Фикс. ставка",
               "Премия", "Штрафы", "Авансы", "Надо выплатить\n(баланс)",
               "Наличными", "Перевод", "Выплачено", "Остаток"]
    HEAD_ROW = 4
    for col, h in enumerate(headers, 1):
        c = _cell(ws, HEAD_ROW, col, h, s, font=s["th_font"], fill=s["th_fill"], center=True)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ---- Данные ----
    FIRST_ROW = HEAD_ROW + 1
    row = FIRST_ROW
    for w in workers_data:
        _cell(ws, row, 1, w['name'], s)

        cell = ws.cell(row=row, column=2, value=w['work_summary'])
        cell.border = s["border"]
        cell.alignment = Alignment(wrap_text=True, vertical='top')

        earned = round(w.get('earned', 0), 2)
        _cell(ws, row, 3, earned or None, s, fmt=money_fmt, center=True)
        _cell(ws, row, 4, (w['fixed_salary'] or 0) or None, s, fmt=money_fmt, center=True)
        _cell(ws, row, 5, (w['bonus'] or 0) or None, s, fmt=money_fmt, center=True)
        _cell(ws, row, 6, (w['penalties'] or 0) or None, s, fmt=money_fmt, center=True)
        _cell(ws, row, 7, (w['advances'] or 0) or None, s, fmt=money_fmt, center=True)

        to_pay = round(w['balance'], 2)
        _cell(ws, row, 8, to_pay, s,
              font=Font(bold=True), fill=green_fill, fmt=money_fmt, center=True)

        # Наличными / Перевод — вручную (жёлтые)
        _cell(ws, row, 9, None, s, fill=input_fill, fmt=money_fmt, center=True)
        _cell(ws, row, 10, None, s, fill=input_fill, fmt=money_fmt, center=True)
        # Выплачено = Наличными + Перевод
        _cell(ws, row, 11, f"=I{row}+J{row}", s, fmt=money_fmt, center=True)
        # Остаток = Надо выплатить − Выплачено
        _cell(ws, row, 12, f"=H{row}-K{row}", s, fmt=money_fmt, center=True)
        row += 1

    LAST_ROW = row - 1  # последняя строка с работником

    # ---- Итоговая строка ----
    total_row = row
    _cell(ws, total_row, 1, "ИТОГО", s, font=tf_font, fill=tf)
    _cell(ws, total_row, 2, "", s, fill=tf)
    if LAST_ROW >= FIRST_ROW:
        for col_letter, col_idx in [('C', 3), ('D', 4), ('E', 5), ('F', 6),
                                    ('G', 7), ('H', 8), ('I', 9), ('J', 10),
                                    ('K', 11), ('L', 12)]:
            _cell(ws, total_row, col_idx,
                  f"=SUM({col_letter}{FIRST_ROW}:{col_letter}{LAST_ROW})",
                  s, font=tf_font, fill=tf, fmt=money_fmt, center=True)
    else:
        for col_idx in range(3, 13):
            _cell(ws, total_row, col_idx, None, s, font=tf_font, fill=tf, fmt=money_fmt, center=True)

    # ==================== БЛОК РАСХОДОВ ====================
    exp_header = total_row + 2
    ws.merge_cells(start_row=exp_header, start_column=1, end_row=exp_header, end_column=7)
    _cell(ws, exp_header, 1, "РАСХОДЫ (свет, аренда, материалы и т.д.)", s,
          font=s["th_font"], fill=s["th_fill"])
    for col_idx in range(2, 8):
        ws.cell(row=exp_header, column=col_idx).fill = s["th_fill"]
    _cell(ws, exp_header, 8, "Сумма", s, font=s["th_font"], fill=s["th_fill"], center=True)

    EXP_ROWS = 10
    exp_first = exp_header + 1
    exp_last = exp_first + EXP_ROWS - 1
    for r in range(exp_first, exp_last + 1):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        _cell(ws, r, 1, None, s, fill=input_fill)          # название — вручную
        for col_idx in range(2, 8):
            ws.cell(row=r, column=col_idx).fill = input_fill
            ws.cell(row=r, column=col_idx).border = s["border"]
        _cell(ws, r, 8, None, s, fill=input_fill, fmt=money_fmt, center=True)  # сумма — вручную

    exp_total = exp_last + 1
    ws.merge_cells(start_row=exp_total, start_column=1, end_row=exp_total, end_column=7)
    _cell(ws, exp_total, 1, "Итого расходов", s, font=tf_font, fill=tf)
    for col_idx in range(2, 8):
        ws.cell(row=exp_total, column=col_idx).fill = tf
    _cell(ws, exp_total, 8, f"=SUM(H{exp_first}:H{exp_last})", s,
          font=tf_font, fill=tf, fmt=money_fmt, center=True)

    # ==================== СВЕРКА КАССЫ ====================
    k_header = exp_total + 2
    ws.merge_cells(start_row=k_header, start_column=1, end_row=k_header, end_column=7)
    _cell(ws, k_header, 1, "СВЕРКА КАССЫ", s, font=s["th_font"], fill=s["th_fill"])
    for col_idx in range(2, 8):
        ws.cell(row=k_header, column=col_idx).fill = s["th_fill"]
    _cell(ws, k_header, 8, "Сумма", s, font=s["th_font"], fill=s["th_fill"], center=True)

    def _kassa_row(r, label, value, *, fill=None, font=None, is_formula=False):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        _cell(ws, r, 1, label, s, font=font, fill=fill)
        for col_idx in range(2, 8):
            if fill:
                ws.cell(row=r, column=col_idx).fill = fill
            ws.cell(row=r, column=col_idx).border = s["border"]
        _cell(ws, r, 8, value, s, font=font, fill=fill, fmt=money_fmt, center=True)

    r = k_header + 1
    got_row = r
    _kassa_row(r, "Получено от директора", None, fill=input_fill, font=Font(bold=True))  # вручную
    r += 1
    salary_row = r
    _kassa_row(r, "Итого зарплата (надо выплатить)", f"=H{total_row}")
    r += 1
    exp_ref_row = r
    _kassa_row(r, "Итого расходы", f"=H{exp_total}")
    r += 1
    need_row = r
    _kassa_row(r, "Всего к выплате (зарплата + расходы)", f"=H{salary_row}+H{exp_ref_row}",
               fill=tf, font=tf_font)
    r += 1
    _kassa_row(r, "   в т.ч. выдано наличными", f"=I{total_row}")
    r += 1
    _kassa_row(r, "   в т.ч. выдано переводом", f"=J{total_row}")
    r += 1
    diff_row = r
    _kassa_row(r, "Разница (получено − всего к выплате)",
               f"=H{got_row}-H{need_row}", fill=warn_fill, font=Font(bold=True, size=11))

    # ---- Ширина столбцов ----
    for col, w in zip('ABCDEFGHIJKL',
                      [22, 42, 13, 13, 12, 12, 12, 15, 13, 13, 13, 13]):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[HEAD_ROW].height = 34

    filename = f"vedomost_{year}_{month:02d}.xlsx"
    wb.save(filename)
    return filename